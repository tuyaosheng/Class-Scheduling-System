"""命令行入口。

  python -m scheduler.cli import 任课与排课说明.xlsx --write
"""
import argparse
import sys
from collections import OrderedDict
from pathlib import Path

from .core import calendar as cal
from .core.config import load_config
from .core.importer import (
    COLUMNS, FIRST_DATA_ROW, import_excel, write_rules_yaml, write_teaching_yaml,
)
from .core.ruletext import parse_fixed_slots, parse_remark, parse_requirement, parse_time_expr

DEFAULT_CONFIG_DIR = Path(__file__).resolve().parent / 'config'


def _fmt_slots(slots):
    """{(0,4),(0,5),(2,1)} -> '周一 4,5 | 周三 1'"""
    by_day = OrderedDict()
    for day, period in sorted(slots):
        by_day.setdefault(cal.DAYS[day], []).append(period)
    return ' | '.join('%s %s' % (d, ','.join(str(p) for p in ps))
                      for d, ps in by_day.items())


def _echo_column(rows, column, parser):
    """按原文去重，输出「原文 -> 解析结果」对照。"""
    lines = []
    seen = set()
    for row in rows:
        raw = row[COLUMNS[column]]
        raw = '' if raw is None else str(raw).strip()
        if not raw or raw in seen:
            continue
        seen.add(raw)
        parsed = parser(raw)
        if isinstance(parsed, set):
            rendered = _fmt_slots(parsed)
        else:
            rendered = '; '.join('%s %s' % (f['type'], f['params']) for f in parsed)
        lines.append('  %s -> %s' % (raw, rendered))
    return lines


def render_import_report(result, rows=None) -> str:
    ds = result.dataset
    used = {}
    for task in ds.tasks:
        if task.consumes_slot:
            used[task.class_id] = used.get(task.class_id, 0) + task.periods
    occupancy = sorted(set(used.values()))

    out = [
        '=== 导入概览（%s）===' % ds.grade,
        '教师 %d 人 | 班级 %d 个 | 任务 %d 个（占格 %d 个）'
        % (len(ds.teachers), len(ds.classes), len(ds.tasks),
           sum(1 for t in ds.tasks if t.consumes_slot)),
        '每班占格数 %s / 每周 %d 格' % (occupancy, cal.N_SLOTS),
        '生成规则 %d 条' % len(result.rules),
        '',
    ]
    if rows is not None:
        out.append('=== 中文规则解析回显（请逐条核对）===')
        for column, parser in [('不能排课节次', parse_time_expr),
                               ('固定节次', parse_fixed_slots),
                               ('排课要求', parse_requirement),
                               ('备注', parse_remark)]:
            out.append('[%s]' % column)
            out.extend(_echo_column(rows, column, parser))
            out.append('')
    if result.warnings:
        out.append('=== 警告 ===')
        out.extend('  ' + w for w in result.warnings)
    else:
        out.append('无警告。')
    return '\n'.join(out)


def _read_rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    return [r for r in wb[wb.sheetnames[0]].iter_rows(min_row=FIRST_DATA_ROW, values_only=True)
            if r and r[COLUMNS['姓名']]]


def cmd_import(args) -> int:
    cfg = load_config(args.config_dir)
    result = import_excel(args.excel, cfg, grade=args.grade)
    print(render_import_report(result, rows=_read_rows(args.excel)))
    if args.write:
        out_dir = Path(args.out_dir or args.config_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        write_teaching_yaml(result, out_dir / 'teaching.yaml')
        write_rules_yaml(result, out_dir / 'rules.generated.yaml')
        print('\n已写入 %s' % out_dir)
    else:
        print('\n（未加 --write，本次只回显不落盘）')
    return 0


def _suffixed_path(out, index):
    out = Path(out)
    return str(out.with_name('%s_候选%d%s' % (out.stem, index, out.suffix)))


def _template_suffixed_path(out, index, count):
    out = Path(out)
    suffix = '_候选%d_模板' % index if count > 1 else '_模板'
    return str(out.with_name('%s%s%s' % (out.stem, suffix, out.suffix)))


def cmd_solve(args) -> int:
    from .core.diagnose import format_conflict, minimal_conflict
    from .core.exporter import export_excel, export_to_template
    from .core.importer import import_excel
    from .core.precheck import format_issues, precheck
    from .core.rules import load_rules
    from .core.solver import solve_many
    from .core.verifier import format_violations, verify, verify_soft

    config_dir = Path(args.config_dir)
    cfg = load_config(config_dir)
    result = import_excel(args.excel, cfg, grade=args.grade)
    rules = load_rules(config_dir / 'rules.yaml', config_dir / 'rules.generated.yaml')

    # L1：总是先跑，毫秒级
    issues = precheck(result.dataset, cfg, rules)
    print(format_issues(issues))
    if issues:
        print('\n预检未通过，不进入求解器。请先处理上述问题。')
        return 2

    solutions = solve_many(result.dataset, cfg, rules, count=args.count,
                           min_diff=args.min_diff, max_seconds=args.max_seconds)

    if not solutions:
        # L2：预检通过却无解，取最小冲突集
        print('\n状态 INFEASIBLE')
        print('\n' + format_conflict(
            minimal_conflict(result.dataset, cfg, rules, max_seconds=args.max_seconds)))
        return 1

    if len(solutions) < args.count:
        print('\n只求出 %d 个彼此有差异（≥%d 处不同）的解，不足要求的 %d 个——'
              '差异空间已经用尽，不是求解失败。' % (len(solutions), args.min_diff, args.count))

    paths = [args.out] if args.count == 1 else [_suffixed_path(args.out, i + 1)
                                                 for i in range(len(solutions))]
    for i, (solution, path) in enumerate(zip(solutions, paths), start=1):
        print('\n[方案 %d] 状态 %s，耗时 %.2f 秒，放置 %d 节课'
              % (i, solution.status, solution.wall_time, len(solution.placements)))
        print(format_violations(verify(solution, result.dataset, cfg, rules)))
        soft = verify_soft(solution, result.dataset, cfg, rules)
        if soft:
            print('  （软偏好：%d 处未完全满足，见上「教师半天连堂过长」）' % len(soft))
        elif any(r.mode == 'soft' and r.enabled for r in rules):
            print('  （软偏好全部满足）')
        export_excel(solution, result.dataset, path, cfg=cfg)
        print('已导出 %s' % path)
        if args.template:
            template_path = _template_suffixed_path(args.out, i, len(solutions))
            export_to_template(solution, result.dataset, args.template, template_path,
                               sheet_name=args.template_sheet, cfg=cfg)
            print('已按模板导出 %s' % template_path)
    return 0


def main(argv=None) -> int:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    parser = argparse.ArgumentParser(prog='scheduler')
    sub = parser.add_subparsers(dest='command', required=True)

    p = sub.add_parser('import', help='导入 Excel 任课表并回显解析结果')
    p.add_argument('excel')
    p.add_argument('--grade', default='初三')
    p.add_argument('--config-dir', default=str(DEFAULT_CONFIG_DIR))
    p.add_argument('--out-dir', default=None)
    p.add_argument('--write', action='store_true', help='确认无误后写入 YAML')
    p.set_defaults(func=cmd_import)

    p = sub.add_parser('solve', help='求解课表并导出 Excel')
    p.add_argument('excel', nargs='?', default='任课与排课说明.xlsx')
    p.add_argument('--grade', default='初三')
    p.add_argument('--config-dir', default=str(DEFAULT_CONFIG_DIR))
    p.add_argument('--out', default='output/课表.xlsx')
    p.add_argument('--max-seconds', type=int, default=60)
    p.add_argument('--count', type=int, default=1,
                   help='生成几个彼此有差异的候选方案（>1 时文件名自动加「_候选N」后缀）')
    p.add_argument('--min-diff', type=int, default=8,
                   help='候选方案之间至少相差几处排课（--count > 1 时才有意义）')
    p.add_argument('--template', default=None,
                   help='按此 Excel 模板版式额外导出一份（如 课程表模板.xlsx）')
    p.add_argument('--template-sheet', default='下学期',
                   help='模板里要写入的工作表名，默认「下学期」')
    p.set_defaults(func=cmd_solve)

    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == '__main__':
    raise SystemExit(main())
