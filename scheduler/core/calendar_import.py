"""作息表批量导入：从 Excel 解析每个年级的节次、钟点、午休边界。

`作息表模板.xlsx` 一个 sheet 一个年级，每行是"第 N 节 / 时间段"（如
"8:25－9:05"）。sheet 名不要求跟系统里的年级名一致——由调用方（API 层）
让用户手动选"这个 sheet 对应哪个年级"，这里只管解析，不管映射。

午休边界用相邻两节课起始时间的最大间隔来推断：一天里唯一的长间隔就是
午休，缺口前最后一节课就是 midday_break_after。这个推断在七年级（8 节，
午休在第 4 节后）和八/九年级（9 节，午休在第 5 节后）两种真实作息上都验证过。
"""
import re
from typing import List, Tuple

import openpyxl
from pydantic import BaseModel

_TIME_RE = re.compile(r'^(\d{1,2}):(\d{2})$')
_DASH_RE = re.compile(r'[－\-~]')


class CalendarParseError(ValueError):
    """作息表格式无法解析。"""


class ParsedCalendarSheet(BaseModel):
    sheet_name: str
    periods_per_day: int
    midday_break_after: int
    clock_times: List[Tuple[str, str]]


def _parse_time_range(text) -> Tuple[str, str]:
    parts = _DASH_RE.split(str(text).strip(), maxsplit=1)
    if len(parts) != 2:
        raise CalendarParseError('无法解析时间段：%r（应形如 8:25－9:05）' % text)
    start, end = parts[0].strip(), parts[1].strip()
    if not _TIME_RE.match(start) or not _TIME_RE.match(end):
        raise CalendarParseError('时间格式不对：%r' % text)
    return start, end


def _minutes(hm: str) -> int:
    h, m = _TIME_RE.match(hm).groups()
    return int(h) * 60 + int(m)


def _infer_midday_break(clock_times: List[Tuple[str, str]]) -> int:
    """相邻两节课起始时间间隔最大的那一处即午休——返回缺口前最后一节的节次号（1-based）。"""
    starts = [_minutes(s) for s, _ in clock_times]
    gaps = [starts[i + 1] - starts[i] for i in range(len(starts) - 1)]
    if not gaps:
        return len(clock_times)
    return gaps.index(max(gaps)) + 1


def parse_calendar_workbook(path) -> List[ParsedCalendarSheet]:
    """逐 sheet 解析，跳过没有任何有效时间行的 sheet（比如空白说明页）。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [r for r in ws.iter_rows(min_row=1, values_only=True) if r and r[0] and len(r) > 1 and r[1]]
        if not rows:
            continue
        clock_times = [_parse_time_range(r[1]) for r in rows]
        out.append(ParsedCalendarSheet(
            sheet_name=sheet_name,
            periods_per_day=len(clock_times),
            midday_break_after=_infer_midday_break(clock_times),
            clock_times=clock_times,
        ))
    return out
