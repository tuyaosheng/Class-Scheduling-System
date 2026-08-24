"""课表导出为 Excel。"""
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font


_RESERVED_LABEL = '（教务固定安排）'


def _write_grid(ws, columns, cell_text, calendar, reserved_slots=()):
    ws.cell(row=1, column=1, value='星期').font = Font(bold=True)
    ws.cell(row=1, column=2, value='节次').font = Font(bold=True)
    for col, name in enumerate(columns, start=3):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    for slot in range(calendar.n_slots):
        day, period = calendar.slot_of(slot)
        row = slot + 2
        ws.cell(row=row, column=1, value=calendar.days[day])
        ws.cell(row=row, column=2, value=period)
        for col, name in enumerate(columns, start=3):
            text = cell_text(name, slot) or (_RESERVED_LABEL if slot in reserved_slots else '')
            if text:
                cell = ws.cell(row=row, column=col, value=text)
                cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'C2'


def export_excel(solution, dataset, path, cfg=None) -> None:
    by_class = defaultdict(list)
    by_teacher = defaultdict(list)
    for p in solution.placements:
        by_class[(p.class_id, p.slot)].append(p)
        by_teacher[(p.teacher, p.slot)].append(p)

    reserved = cfg.reserved_slot_indices(dataset.grade) if cfg else set()

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = '班级课表'
    _write_grid(
        ws, dataset.classes,
        lambda class_id, slot: '/'.join(
            '%s%s' % (p.course, '(%s)' % p.parity if p.parity else '')
            for p in by_class.get((class_id, slot), [])),
        dataset.calendar,
        reserved_slots=reserved)

    ws2 = wb.create_sheet('教师课表')
    _write_grid(
        ws2, sorted(dataset.teachers),
        lambda teacher, slot: '/'.join(
            '%d班%s' % (p.class_id, p.course)
            for p in by_teacher.get((teacher, slot), [])),
        dataset.calendar)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------- 按教务模板导出

# 「课程表模板.xlsx」『下学期』工作表的版式：行是班级，列是星期×节次。
# 初三1 在第 12 行，此后每班占一行；每天 9 列（节次 1-8 + 第二课堂=第9节），
# 周一从第 2 列开始。这两个偏移量是模板本身的版式，不是系统的计算结果。
_TEMPLATE_FIRST_CLASS_ROW = 12
_TEMPLATE_FIRST_DAY_COL = 2


def _template_cell(class_id, slot, calendar):
    day, period = calendar.slot_of(slot)
    row = _TEMPLATE_FIRST_CLASS_ROW + (class_id - 1)
    col = _TEMPLATE_FIRST_DAY_COL + day * calendar.periods_per_day + (period - 1)
    return row, col


_PARITY_ORDER = {'单周': 0, '双周': 1}
_TEMPLATE_NOTE_HEADER_ROW = 11   # 与「节  次」表头同一行


def _cell_text(placements, cfg):
    """单双周家族（心美）整格折叠成家族名；其余按原样拼课程名。"""
    if cfg is not None and placements:
        families = {cfg.courses[p.course].family for p in placements
                    if cfg.courses[p.course].alternate}
        if len(families) == 1 and all(cfg.courses[p.course].alternate for p in placements):
            return next(iter(families))
    return '/'.join('%s%s' % (p.course, '(%s)' % p.parity if p.parity else '')
                    for p in placements)


def _alternate_note(placements_for_class_family):
    ordered = sorted(placements_for_class_family,
                     key=lambda p: _PARITY_ORDER.get(p.parity, 99))
    return '/'.join('%s(%s)' % (p.course, p.parity) for p in ordered)


def export_to_template(solution, dataset, template_path, out_path, sheet_name='下学期',
                       cfg=None) -> None:
    """按教务提供的『课程表模板.xlsx』版式导出。

    模板里教务固定占位的 8 格（班会/体比/校本1/综实2/体选）已经预填好内容，
    这里只写系统求解出的格子，不碰模板已有的表头、预填内容与会议安排说明。

    传入 cfg 时，单双周家族（心美）在格子里只显示家族名，具体单双周安排改写到
    最后一列（每个家族一列），避免格子里塞下两门课的名字挤占版面。
    """
    wb = openpyxl.load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError('模板 %s 里没有名为 %r 的工作表' % (template_path, sheet_name))
    ws = wb[sheet_name]

    by_cell = defaultdict(list)
    for p in solution.placements:
        by_cell[_template_cell(p.class_id, p.slot, dataset.calendar)].append(p)

    for (row, col), placements in by_cell.items():
        ws.cell(row=row, column=col, value=_cell_text(placements, cfg))

    if cfg is not None:
        by_class_family = defaultdict(list)
        for p in solution.placements:
            course = cfg.courses[p.course]
            if course.alternate:
                by_class_family[(p.class_id, course.family)].append(p)
        families = sorted({family for _, family in by_class_family})
        note_col_of = {family: _TEMPLATE_FIRST_DAY_COL + len(dataset.calendar.days) * dataset.calendar.periods_per_day + i
                      for i, family in enumerate(families)}
        for family, col in note_col_of.items():
            ws.cell(row=_TEMPLATE_NOTE_HEADER_ROW, column=col, value='%s单双周安排' % family)
        for (class_id, family), placements in by_class_family.items():
            row = _TEMPLATE_FIRST_CLASS_ROW + (class_id - 1)
            ws.cell(row=row, column=note_col_of[family], value=_alternate_note(placements))

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
