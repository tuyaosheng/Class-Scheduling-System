"""Excel 任课表导入。

关键口径（见计划文档「三处口径」）：
  1. 「不能排课节次」按教师取并集 —— 空白行不代表该课可排在会议时段。
  2. 「固定节次」是窗口不是钉死。
  3. 「周课时 0.5」转成 1 节 + 周次奇偶标记。
"""
import json
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import yaml
from pydantic import BaseModel

from .models import Dataset, Teacher, TeachingTask
from .ruletext import parse_fixed_slots, parse_remark, parse_requirement, parse_time_expr

COLUMNS = {
    '姓名': 0, '任教年级': 1, '学科': 2, '任教班': 3, '周课时': 4,
    '职务': 5, '固定节次': 6, '不能排课节次': 7, '排课要求': 8, '备注': 9,
}
FIRST_DATA_ROW = 3

# 规则片段的 scope 维度：daily_* 按学科系统计，连堂/间隔按具体课程
_FAMILY_SCOPED = {'daily_min', 'daily_max', 'weekday_exact'}
_COURSE_SCOPED = {'consecutive', 'spacing'}


class ImportResult(BaseModel):
    dataset: Dataset
    rules: List[dict]
    warnings: List[str]
    conflicts: List[dict] = []
    rule_echo: Dict[str, List[dict]] = {}


def _cell(row, name):
    value = row[COLUMNS[name]]
    return '' if value is None else str(value).strip()


def _class_ids(text):
    return [int(x) for x in text.split(',') if x.strip()]


def import_excel(path, cfg, grade='初三') -> ImportResult:
    calendar = cfg.calendar_of(grade)
    courses = cfg.courses_of(grade)
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = [r for r in wb[wb.sheetnames[0]].iter_rows(min_row=FIRST_DATA_ROW, values_only=True)
            if r and r[COLUMNS['姓名']]]

    # ---- 第一遍：教师级信息聚合 ----
    forbidden = defaultdict(set)
    duties = defaultdict(set)
    for row in rows:
        name = _cell(row, '姓名')
        forbidden[name] |= parse_time_expr(_cell(row, '不能排课节次'), calendar)
        for duty in _cell(row, '职务').split(','):
            if duty.strip():
                duties[name].add(duty.strip())

    teachers = {
        name: Teacher(name=name,
                      duties=sorted(duties[name]),
                      forbidden=sorted([d, p] for d, p in forbidden[name]))
        for name in forbidden
    }

    # ---- 第二遍：教学任务 ----
    tasks: List[TeachingTask] = []
    classes = set()
    for row in rows:
        name, course = _cell(row, '姓名'), _cell(row, '学科')
        if course not in courses:
            raise ValueError('Excel 中的学科 %r 不在课程目录里' % course)
        if courses[course].external:
            for class_id in _class_ids(_cell(row, '任教班')):
                classes.add(class_id)
            continue
        hours = float(_cell(row, '周课时'))
        base_parity = None
        if hours == 0.5:
            base_parity = courses[course].alternate
            if not base_parity:
                raise ValueError('%s 周课时 0.5 但课程目录未声明 alternate' % course)
            periods = 1
        else:
            if hours != int(hours):
                raise ValueError('%s %s 周课时 %s 既不是整数也不是 0.5' % (name, course, hours))
            periods = int(hours)
        for class_id in _class_ids(_cell(row, '任教班')):
            classes.add(class_id)
            parity = base_parity
            if base_parity and class_id % 2 == 0:
                # 按班号奇偶各半翻转单双周，否则该老师整学期只在单周（或双周）
                # 有课，负荷忽高忽低；翻转后每周教的班数固定，见坑 3。
                parity = '双周' if base_parity == '单周' else '单周'
            tasks.append(TeachingTask(id=len(tasks), grade=grade, class_id=class_id,
                                      course=course, teacher=name,
                                      periods=periods, parity=parity))

    dataset = Dataset(grade=grade, classes=sorted(classes),
                      teachers=teachers, tasks=tasks, calendar=calendar)
    rules = _build_rules(rows, cfg, grade, forbidden, calendar)
    warnings = _check_class_loads(dataset, cfg, grade)
    return ImportResult(dataset=dataset, rules=rules, warnings=warnings)


def _build_rules(rows, cfg, grade, forbidden, calendar) -> List[dict]:
    """forbidden 由调用方（import_excel 的第一遍）算好传入 ——

    教师禁排并集只应算一次；这里不再重新遍历 rows 推导它，
    避免与 Teacher.forbidden 各自独立计算导致悄悄分叉（见 review Important）。
    """
    rules: List[dict] = []
    courses = cfg.courses_of(grade)

    # 教师禁排：来自 import_excel 已聚合好的并集
    for name in sorted(forbidden):
        if not forbidden[name]:
            continue
        rules.append({
            'type': 'forbid_slots',
            'scope': {'grade': grade, 'teacher': name},
            'params': {'slots': sorted([d, p] for d, p in forbidden[name])},
            'mode': 'hard',
        })

    # 固定节次：按课程去重。教务已固定安排的课程（external）不生成任务，
    # 也就不需要 pin_window——它们的窗口改由 reserved_slots 统一挖空处理。
    pins = defaultdict(set)
    for row in rows:
        course = _cell(row, '学科')
        if courses[course].external:
            continue
        slots = parse_fixed_slots(_cell(row, '固定节次'), calendar)
        if slots:
            pins[course] |= slots
    for course in sorted(pins):
        rules.append({
            'type': 'pin_window',
            'scope': {'grade': grade, 'course': course},
            'params': {'slots': sorted([d, p] for d, p in pins[course])},
            'mode': 'hard',
        })

    # 教务固定占位的时段：整体从求解器可用格位里挖空，禁止任何常规任务占用
    reserved = cfg.reserved_slots.get(grade) or []
    if reserved:
        rules.append({
            'type': 'forbid_slots',
            'scope': {'grade': grade},
            'params': {'slots': sorted([d, p] for d, p in reserved)},
            'mode': 'hard',
        })

    # 排课要求与备注：按 (类型, 作用域键, 参数) 去重
    seen = set()
    for row in rows:
        course = _cell(row, '学科')
        fragments = (parse_requirement(_cell(row, '排课要求'))
                     + parse_remark(_cell(row, '备注')))
        for frag in fragments:
            rtype, params = frag['type'], frag['params']
            if rtype in _FAMILY_SCOPED:
                scope = {'grade': grade, 'family': cfg.family_of(grade, course)}
            elif rtype in _COURSE_SCOPED:
                scope = {'grade': grade, 'course': course}
            else:                       # alternate_weeks 等年级级规则
                scope = {'grade': grade}
                params = {k: v for k, v in params.items() if k != 'self_parity'}
            key = (rtype, tuple(sorted(scope.items())),
                   yaml.safe_dump(params, sort_keys=True, allow_unicode=True))
            if key in seen:
                continue
            seen.add(key)
            rules.append({'type': rtype, 'scope': scope, 'params': params,
                          'mode': 'hard' if rtype != 'spacing' else 'soft'})

    for rule in rules:
        if rule['mode'] == 'soft':
            rule.setdefault('enabled', True)
            rule.setdefault('weight', 5)
    return rules


def _resolve_course_periods(cfg, grade) -> Dict[str, int]:
    """把课程计划的键（可能是学科系名，比如"心美"代表美术+心理各占 1 节）
    展开成"课程名 -> 每周节数"的映射，供纯任课表导入（不经过排课说明.xlsx）
    确定每个任务的 periods 用。"""
    plan = cfg.plans.get(grade) or {}
    out: Dict[str, int] = {}
    for key, hours in plan.items():
        for name in cfg.resolve_plan_key(grade, key):
            out[name] = hours
    return out


def build_dataset_from_pivot(pivot: Dict[Tuple[int, str], str], cfg, grade='初三',
                             existing_teachers: Dict[str, Teacher] = None) -> ImportResult:
    """任课表（班别×学科矩阵）是"谁教谁"的唯一来源——不依赖排课说明.xlsx
    提供班级/课程/周课时，那些信息分别来自课程与学科系步骤的课程计划（子项目2）
    和年级日历。排课说明.xlsx 降级为纯规则文本表，在另一条导入路径里按教师
    姓名匹配（子项目5），不在这里处理，所以这里产出的 rules 恒为空列表——
    规则由 rules.yaml/rules.generated.yaml 在求解时另行加载，不是这次导入的一部分。

    pivot 由 parse_teaching_table 解析得到，也可以是前端"编辑任课表"页面
    提交的整份矩阵（增删几个格子之后的版本）——两条路径共用同一个构建函数，
    保证「导入」和「编辑保存」产出的 Dataset 形状完全一致。

    existing_teachers：调用方传入"当前已保存的教师信息"（duties/forbidden），
    同名教师直接沿用，而不是每次都建一个空白 Teacher——这条导入路径本身不
    产生禁排/职务信息（那是排课说明.xlsx 的职责，子项目5），如果不做这层
    保留，任何一次"编辑任课表并保存"都会把之前排课说明导入算出来的教师
    禁排/职务信息整体清空——这是真实发生过的数据丢失（浏览器实测中招过一次）。
    """
    existing_teachers = existing_teachers or {}
    calendar = cfg.calendar_of(grade)
    courses = cfg.courses_of(grade)
    course_periods = _resolve_course_periods(cfg, grade)

    tasks: List[TeachingTask] = []
    classes = set()
    teacher_names = set()
    for task_id, ((class_id, course), teacher) in enumerate(sorted(pivot.items())):
        classes.add(class_id)
        if course not in courses:
            raise ValueError('任课表中的学科 %r 不在 %s 的课程目录里' % (course, grade))
        if courses[course].external:
            continue   # 教务固定占位的课程不生成任务，任课信息不进 Dataset
        if course not in course_periods:
            raise ValueError('课程 %r 没有在 %s 的课程计划里设置周课时' % (course, grade))
        teacher_names.add(teacher)
        base_parity = courses[course].alternate
        parity = base_parity
        if base_parity and class_id % 2 == 0:
            # 按班号奇偶各半翻转单双周，否则该老师整学期只在单周（或双周）
            # 有课，负荷忽高忽低；翻转后每周教的班数固定，见坑 3。
            parity = '双周' if base_parity == '单周' else '单周'
        tasks.append(TeachingTask(id=task_id, grade=grade, class_id=class_id,
                                  course=course, teacher=teacher,
                                  periods=course_periods[course], parity=parity))

    teachers = {
        name: existing_teachers[name] if name in existing_teachers else Teacher(name=name)
        for name in teacher_names
    }
    dataset = Dataset(grade=grade, classes=sorted(classes), teachers=teachers,
                      tasks=tasks, calendar=calendar)
    warnings = _check_class_loads(dataset, cfg, grade)
    return ImportResult(dataset=dataset, rules=[], warnings=warnings)


def import_teaching_table(path, cfg, grade='初三', existing_teachers: Dict[str, Teacher] = None) -> ImportResult:
    pivot = parse_teaching_table(path, cfg, grade)
    return build_dataset_from_pivot(pivot, cfg, grade, existing_teachers=existing_teachers)


def _check_class_loads(dataset, cfg, grade) -> List[str]:
    plan_total = sum((cfg.plans.get(grade) or {}).values())
    used = defaultdict(int)
    for task in dataset.tasks:
        if task.consumes_slot:
            used[task.class_id] += task.periods
    warnings = []
    for class_id in sorted(used):
        if plan_total and used[class_id] != plan_total:
            warnings.append('%d班占 %d 格，课程计划为 %d 格'
                            % (class_id, used[class_id], plan_total))
    return warnings


def write_teaching_yaml(result: ImportResult, path) -> None:
    data = {
        'grade': result.dataset.grade,
        'classes': result.dataset.classes,
        'teachers': [t.model_dump() for t in result.dataset.teachers.values()],
        'tasks': [t.model_dump() for t in result.dataset.tasks],
    }
    Path(path).write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding='utf-8')


def write_rules_yaml(result: ImportResult, path) -> None:
    Path(path).write_text(
        yaml.safe_dump({'rules': result.rules}, allow_unicode=True, sort_keys=False),
        encoding='utf-8')


TEACHING_TABLE_HEADER_ROW = 2
TEACHING_TABLE_FIRST_DATA_ROW = 3


def parse_teaching_table(path, cfg, grade='初三') -> Dict[Tuple[int, str], str]:
    """解析『班别 × 学科』矩阵版式的任课表，返回 (班级,课程) -> 教师。"""
    course_catalog = cfg.courses_of(grade)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[TEACHING_TABLE_HEADER_ROW]]
    courses = header[1:]
    for course in courses:
        if course and course not in course_catalog:
            raise ValueError("任课表中的学科 %r 不在课程目录里" % course)

    pivot: Dict[Tuple[int, str], str] = {}
    for row in ws.iter_rows(min_row=TEACHING_TABLE_FIRST_DATA_ROW, values_only=True):
        if not row or row[0] is None:
            continue
        class_id = int(row[0])
        for course, teacher in zip(courses, row[1:]):
            if course and teacher:
                pivot[(class_id, course)] = str(teacher).strip()
    return pivot


def _read_rule_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return [r for r in wb[wb.sheetnames[0]].iter_rows(min_row=FIRST_DATA_ROW, values_only=True)
            if r and r[COLUMNS['姓名']]]


def _fmt_slots_display(slots, calendar):
    by_day = {}
    for d, p in sorted(slots):
        by_day.setdefault(calendar.days[d], []).append(p)
    return ' | '.join('%s %s' % (day, ','.join(str(p) for p in ps))
                      for day, ps in by_day.items())


def merge_teaching_and_rules(teaching_path, rules_path, cfg, grade='初三',
                             rule_engine='regex', ai_client=None) -> ImportResult:
    """两源合并：任课表给『谁教谁』，排课说明给『周课时与规则文本』。

    (班级,课程) 两边教师对不上、或排课说明里缺对应行，一律计入 conflicts，
    不生成任务——不静默二选一（见 spec 5.2）。
    """
    from .ruletext import parse_fixed_slots, parse_remark, parse_requirement, parse_time_expr

    calendar = cfg.calendar_of(grade)
    courses = cfg.courses_of(grade)
    teaching_pivot = parse_teaching_table(teaching_path, cfg, grade)
    rows = _read_rule_rows(rules_path)

    forbidden: Dict[str, set] = defaultdict(set)
    duties: Dict[str, set] = defaultdict(set)
    for row in rows:
        name = _cell(row, '姓名')
        for duty in _cell(row, '职务').split(','):
            if duty.strip():
                duties[name].add(duty.strip())
        if rule_engine != 'ai':
            # AI 模式下禁排改由主循环里逐行解析结果累积（见下方 parsed.not_available），
            # 不在这里用正则重复算一遍——否则 AI 对这一列的解析结果会被静默丢弃。
            forbidden[name] |= parse_time_expr(_cell(row, '不能排课节次'), calendar)

    tasks: List[TeachingTask] = []
    classes = set()
    conflicts: List[dict] = []
    covered_keys = set()

    rule_echo: Dict[str, List[dict]] = {
        '不能排课节次': [], '固定节次': [], '排课要求': [], '备注': [],
    }
    seen_raw: Dict[str, set] = {k: set() for k in rule_echo}

    pins = defaultdict(set)
    rule_seen = set()
    rules: List[dict] = []

    for row in rows:
        rules_teacher, course = _cell(row, '姓名'), _cell(row, '学科')
        if course not in courses:
            raise ValueError('Excel 中的学科 %r 不在课程目录里' % course)

        not_avail_raw = _cell(row, '不能排课节次')
        fixed_raw = _cell(row, '固定节次')
        req_raw = _cell(row, '排课要求')
        remark_raw = _cell(row, '备注')

        if rule_engine == 'ai':
            from scheduler.ai.rule_parser import parse_row_ai
            parsed = parse_row_ai(not_avail_raw, fixed_raw, req_raw, remark_raw, calendar, client=ai_client)
            forbidden[rules_teacher] |= {(d, p) for d, p in parsed.not_available}
            fixed_slots = {(d, p) for d, p in parsed.fixed_slots}
            fragments = parsed.requirement + parsed.remark
        else:
            fixed_slots = parse_fixed_slots(fixed_raw, calendar)
            fragments = parse_requirement(req_raw) + parse_remark(remark_raw)

        if not_avail_raw and not_avail_raw not in seen_raw['不能排课节次']:
            seen_raw['不能排课节次'].add(not_avail_raw)
            not_avail_display_slots = ({(d, p) for d, p in parsed.not_available} if rule_engine == 'ai'
                                       else parse_time_expr(not_avail_raw, calendar))
            rule_echo['不能排课节次'].append(
                {'raw': not_avail_raw, 'parsed': _fmt_slots_display(not_avail_display_slots, calendar)})
        if fixed_raw and fixed_raw not in seen_raw['固定节次']:
            seen_raw['固定节次'].add(fixed_raw)
            rule_echo['固定节次'].append({'raw': fixed_raw, 'parsed': _fmt_slots_display(fixed_slots, calendar)})
        if req_raw and req_raw not in seen_raw['排课要求']:
            seen_raw['排课要求'].add(req_raw)
            rule_echo['排课要求'].append({
                'raw': req_raw,
                'parsed': '; '.join('%s %s' % (f['type'], f['params']) for f in
                                     (parse_requirement(req_raw) if rule_engine != 'ai' else parsed.requirement))})
        if remark_raw and remark_raw not in seen_raw['备注']:
            seen_raw['备注'].add(remark_raw)
            rule_echo['备注'].append({
                'raw': remark_raw,
                'parsed': '; '.join('%s %s' % (f['type'], f['params']) for f in
                                     (parse_remark(remark_raw) if rule_engine != 'ai' else parsed.remark))})

        if not courses[course].external and fixed_slots:
            pins[course] |= fixed_slots

        if not courses[course].external:
            for frag in fragments:
                rtype, params = frag['type'], frag['params']
                if rtype in _FAMILY_SCOPED:
                    scope = {'grade': grade, 'family': cfg.family_of(grade, course)}
                elif rtype in _COURSE_SCOPED:
                    scope = {'grade': grade, 'course': course}
                else:
                    scope = {'grade': grade}
                    params = {k: v for k, v in params.items() if k != 'self_parity'}
                key = (rtype, tuple(sorted(scope.items())), json.dumps(params, sort_keys=True))
                if key in rule_seen:
                    continue
                rule_seen.add(key)
                rules.append({'type': rtype, 'scope': scope, 'params': params,
                              'mode': 'hard' if rtype != 'spacing' else 'soft'})

        if courses[course].external:
            for class_id in _class_ids(_cell(row, '任教班')):
                classes.add(class_id)
                covered_keys.add((class_id, course))
            continue

        hours = float(_cell(row, '周课时'))
        base_parity = None
        if hours == 0.5:
            base_parity = courses[course].alternate
            if not base_parity:
                raise ValueError('%s 周课时 0.5 但课程目录未声明 alternate' % course)
            periods = 1
        else:
            if hours != int(hours):
                raise ValueError('%s %s 周课时 %s 既不是整数也不是 0.5'
                                 % (rules_teacher, course, hours))
            periods = int(hours)

        for class_id in _class_ids(_cell(row, '任教班')):
            classes.add(class_id)
            covered_keys.add((class_id, course))
            key = (class_id, course)
            pivot_teacher = teaching_pivot.get(key)
            if pivot_teacher is None:
                conflicts.append({'class_id': class_id, 'course': course,
                                  'from_teaching_table': None, 'from_rules_sheet': rules_teacher})
                continue
            if pivot_teacher != rules_teacher:
                conflicts.append({'class_id': class_id, 'course': course,
                                  'from_teaching_table': pivot_teacher,
                                  'from_rules_sheet': rules_teacher})
                continue
            parity = base_parity
            if base_parity and class_id % 2 == 0:
                # 按班号奇偶各半翻转单双周，否则该老师整学期只在单周（或双周）
                # 有课，负荷忽高忽低；翻转后每周教的班数固定，见坑 3。
                parity = '双周' if base_parity == '单周' else '单周'
            tasks.append(TeachingTask(id=len(tasks), grade=grade, class_id=class_id,
                                      course=course, teacher=pivot_teacher,
                                      periods=periods, parity=parity))

    for key, teacher in teaching_pivot.items():
        if key not in covered_keys:
            conflicts.append({'class_id': key[0], 'course': key[1],
                              'from_teaching_table': teacher, 'from_rules_sheet': None})

    for name in sorted(forbidden):
        if not forbidden[name]:
            continue
        rules.append({
            'type': 'forbid_slots',
            'scope': {'grade': grade, 'teacher': name},
            'params': {'slots': sorted([d, p] for d, p in forbidden[name])},
            'mode': 'hard',
        })
    for course in sorted(pins):
        rules.append({
            'type': 'pin_window',
            'scope': {'grade': grade, 'course': course},
            'params': {'slots': sorted([d, p] for d, p in pins[course])},
            'mode': 'hard',
        })
    reserved = cfg.reserved_slots.get(grade) or []
    if reserved:
        rules.append({
            'type': 'forbid_slots',
            'scope': {'grade': grade},
            'params': {'slots': sorted([d, p] for d, p in reserved)},
            'mode': 'hard',
        })
    for rule in rules:
        if rule['mode'] == 'soft':
            rule.setdefault('enabled', True)
            rule.setdefault('weight', 5)

    # forbidden 在 AI 模式下要到主循环跑完才最终定型（逐行累积 parsed.not_available），
    # 所以 teachers 字典必须放在主循环之后再建——放前面会在 AI 模式下读到半成品。
    teachers = {
        name: Teacher(name=name, duties=sorted(duties[name]),
                      forbidden=sorted([d, p] for d, p in forbidden[name]))
        for name in forbidden
    }

    dataset = Dataset(grade=grade, classes=sorted(classes), teachers=teachers,
                      tasks=tasks, calendar=calendar)
    warnings = _check_class_loads(dataset, cfg, grade)
    return ImportResult(dataset=dataset, rules=rules, warnings=warnings,
                        conflicts=conflicts, rule_echo=rule_echo)


# ---- 子项目5：排课说明.xlsx 降级为纯规则文本表 ----
# 不再提供任教班/周课时（那是任课表的职责，子项目4），按「姓名+学科」匹配、
# 不需要按班级交叉核对——"谁教谁"已经完全由任课表决定。

RULE_SHEET_COLUMNS = {
    '姓名': 0, '任教年级': 1, '学科': 2, '职务': 3,
    '固定节次': 4, '不能排课节次': 5, '排课要求': 6, '备注': 7,
}
RULE_SHEET_FIRST_DATA_ROW = 3


class RuleImportResult(BaseModel):
    rules: List[dict]
    teacher_facts: List[dict]
    warnings: List[str]
    rule_echo: Dict[str, List[dict]] = {}


def _rule_cell(row, name):
    idx = RULE_SHEET_COLUMNS[name]
    value = row[idx] if idx < len(row) else None
    return '' if value is None else str(value).strip()


def _read_rule_sheet_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return [r for r in wb[wb.sheetnames[0]].iter_rows(min_row=RULE_SHEET_FIRST_DATA_ROW, values_only=True)
            if r and r[RULE_SHEET_COLUMNS['姓名']]]


def _fmt_fragments_display(fragments) -> str:
    return '; '.join('%s %s' % (f['type'], f['params']) for f in fragments)


def import_rule_text_table(path, cfg, grade='初三', ai_client=None) -> RuleImportResult:
    """排课说明.xlsx 降级为纯规则文本表——正则解析永远跑一遍、是真正生效的规则
    来源（铁律5：AI 不做硬性判定）；ai_client 给定时额外跑 AI 复核，只用来发现
    正则漏解析或两者分歧的地方，写进 rule_echo 供人工确认，不单独产出规则。
    AI 复核失败（未配置/网络错误）不影响正则结果，只记一条 warning。
    """
    from .ruletext import parse_fixed_slots, parse_remark, parse_requirement, parse_time_expr

    calendar = cfg.calendar_of(grade)
    courses = cfg.courses_of(grade)
    rows = _read_rule_sheet_rows(path)

    forbidden: Dict[str, set] = defaultdict(set)
    duties: Dict[str, set] = defaultdict(set)
    for row in rows:
        name = _rule_cell(row, '姓名')
        for duty in _rule_cell(row, '职务').split(','):
            if duty.strip():
                duties[name].add(duty.strip())

    rule_echo: Dict[str, List[dict]] = {'不能排课节次': [], '固定节次': [], '排课要求': [], '备注': []}
    seen_raw: Dict[str, set] = {k: set() for k in rule_echo}
    pins = defaultdict(set)
    rule_seen = set()
    rules: List[dict] = []
    warnings: List[str] = []

    def _echo(column, raw, regex_display, ai_display):
        if not raw or raw in seen_raw[column]:
            return
        seen_raw[column].add(raw)
        entry = {'raw': raw, 'parsed': regex_display}
        if ai_display is not None:
            entry['ai_parsed'] = ai_display
            entry['mismatch'] = ai_display != regex_display
        rule_echo[column].append(entry)

    for row in rows:
        name, course = _rule_cell(row, '姓名'), _rule_cell(row, '学科')
        if course not in courses:
            raise ValueError('排课说明中的学科 %r 不在 %s 的课程目录里' % (course, grade))

        not_avail_raw = _rule_cell(row, '不能排课节次')
        fixed_raw = _rule_cell(row, '固定节次')
        req_raw = _rule_cell(row, '排课要求')
        remark_raw = _rule_cell(row, '备注')

        not_avail_slots = parse_time_expr(not_avail_raw, calendar)
        forbidden[name] |= not_avail_slots
        fixed_slots = parse_fixed_slots(fixed_raw, calendar)
        req_fragments = parse_requirement(req_raw)
        remark_fragments = parse_remark(remark_raw)

        ai_parsed = None
        if ai_client is not None and (not_avail_raw or fixed_raw or req_raw or remark_raw):
            from scheduler.ai.rule_parser import parse_row_ai
            try:
                ai_parsed = parse_row_ai(not_avail_raw, fixed_raw, req_raw, remark_raw, calendar, client=ai_client)
            except Exception as exc:
                warnings.append('AI 复核失败（%s %s）：%s；已仅采用正则解析结果' % (name, course, exc))

        _echo('不能排课节次', not_avail_raw, _fmt_slots_display(not_avail_slots, calendar),
              _fmt_slots_display({(d, p) for d, p in ai_parsed.not_available}, calendar) if ai_parsed else None)
        _echo('固定节次', fixed_raw, _fmt_slots_display(fixed_slots, calendar),
              _fmt_slots_display({(d, p) for d, p in ai_parsed.fixed_slots}, calendar) if ai_parsed else None)
        _echo('排课要求', req_raw, _fmt_fragments_display(req_fragments),
              _fmt_fragments_display(ai_parsed.requirement) if ai_parsed else None)
        _echo('备注', remark_raw, _fmt_fragments_display(remark_fragments),
              _fmt_fragments_display(ai_parsed.remark) if ai_parsed else None)

        if not courses[course].external and fixed_slots:
            pins[course] |= fixed_slots

        if not courses[course].external:
            for frag in req_fragments + remark_fragments:
                rtype, params = frag['type'], frag['params']
                if rtype in _FAMILY_SCOPED:
                    scope = {'grade': grade, 'family': cfg.family_of(grade, course)}
                elif rtype in _COURSE_SCOPED:
                    scope = {'grade': grade, 'course': course}
                else:
                    scope = {'grade': grade}
                    params = {k: v for k, v in params.items() if k != 'self_parity'}
                key = (rtype, tuple(sorted(scope.items())), json.dumps(params, sort_keys=True))
                if key in rule_seen:
                    continue
                rule_seen.add(key)
                rules.append({'type': rtype, 'scope': scope, 'params': params,
                              'mode': 'hard' if rtype != 'spacing' else 'soft'})

    for name in sorted(forbidden):
        if not forbidden[name]:
            continue
        rules.append({
            'type': 'forbid_slots',
            'scope': {'grade': grade, 'teacher': name},
            'params': {'slots': sorted([d, p] for d, p in forbidden[name])},
            'mode': 'hard',
        })
    for course in sorted(pins):
        rules.append({
            'type': 'pin_window',
            'scope': {'grade': grade, 'course': course},
            'params': {'slots': sorted([d, p] for d, p in pins[course])},
            'mode': 'hard',
        })
    reserved = cfg.reserved_slots.get(grade) or []
    if reserved:
        rules.append({
            'type': 'forbid_slots',
            'scope': {'grade': grade},
            'params': {'slots': sorted([d, p] for d, p in reserved)},
            'mode': 'hard',
        })
    for rule in rules:
        if rule['mode'] == 'soft':
            rule.setdefault('enabled', True)
            rule.setdefault('weight', 5)

    teacher_facts = [
        {'name': name, 'duties': sorted(duties[name]), 'forbidden': sorted([d, p] for d, p in forbidden[name])}
        for name in sorted(set(duties) | set(forbidden))
    ]

    return RuleImportResult(rules=rules, teacher_facts=teacher_facts, warnings=warnings, rule_echo=rule_echo)


def write_rules_generated_yaml_for_grade(new_rules: List[dict], grade: str, path) -> None:
    """按年级整体替换 rules.generated.yaml 里这个年级的规则，其他年级的规则原样保留。"""
    p = Path(path)
    existing = []
    if p.exists():
        data = yaml.safe_load(p.read_text(encoding='utf-8')) or {}
        existing = data.get('rules') or []
    kept = [r for r in existing if (r.get('scope') or {}).get('grade') != grade]
    p.write_text(yaml.safe_dump({'rules': kept + new_rules}, allow_unicode=True, sort_keys=False),
                encoding='utf-8')


def merge_teacher_facts_into_teaching_yaml(teacher_facts: List[dict], grade: str, path) -> None:
    """把排课说明解析出的教师职务/禁排信息合并进 teaching.yaml——只替换这条
    导入路径提到的教师，没提到的教师（比如任课表已有但排课说明没填规则的）原样
    保留，不清空。要求 teaching.yaml 已经是该年级的数据（先完成任课表导入）。"""
    p = Path(path)
    if not p.exists():
        raise ValueError('请先在「任课表」步骤完成导入，再导入排课规则')
    data = yaml.safe_load(p.read_text(encoding='utf-8')) or {}
    if data.get('grade') != grade:
        raise ValueError('当前任课表是 %r 的数据，和排课规则的年级 %r 不一致'
                         % (data.get('grade'), grade))
    by_name = {t['name']: t for t in data.get('teachers', [])}
    for fact in teacher_facts:
        by_name[fact['name']] = fact
    data['teachers'] = [by_name[name] for name in sorted(by_name)]
    p.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding='utf-8')


# ---- 排课说明.xlsx 填写模板：给教务示范每一类文本该怎么写 ----
# 例子取自 ruletext.py 实测覆盖的真实写法（不能排课节次共 22 种，这里只挑有
# 代表性的几种），不是穷举——穷举的完整清单在「填写说明」sheet 里用文字列出。

_RULE_SHEET_EXAMPLES = [
    # (姓名, 任教年级, 学科, 职务, 固定节次, 不能排课节次, 排课要求, 备注)
    ('示例-张老师', '初三', '语文', '班主任', '', '周二上午不排课', '', ''),
    ('示例-李老师', '初三', '数学', '', '', '周一4、5节，周三上午不排课', '', ''),
    ('示例-王老师', '初三', '英语', '', '', '周五第4，5节不排课', '', ''),
    ('示例-赵老师', '初三', '体育', '', '周一第9节', '', '', ''),
    ('示例-钱老师', '初三', '体育', '', '周二第8、9节', '', '', ''),
    ('示例-孙老师', '初三', '物理', '', '', '', '保证每天有1节', ''),
    ('示例-周老师', '初三', '英语', '', '', '', '同一个班当天不能排2节', ''),
    ('示例-吴老师', '初三', '化学', '', '', '', '保证周一三四每天1节', ''),
    ('示例-郑老师', '初三', '美术', '', '', '', '与心理课分单双周上，即"心美"周课时1节', ''),
    ('示例-冯老师', '初三', '心理', '', '', '', '与美术课分单双周上，即"心美"周课时1节', ''),
    ('示例-陈老师', '初三', '数学', '', '', '', '', '其中一天为连堂课'),
    ('示例-褚老师', '初三', '语文', '', '', '', '', '两个班之间要隔开1节'),
]

_RULE_SHEET_NOTES = [
    '这是排课规则文本的填写示例，不是可以直接导入的真实数据——正式导入前请把',
    '「示例」开头的行整体删除，换成你们学校的真实教师和文本。',
    '',
    '姓名 / 任教年级 / 学科：与「任课表」步骤里的教师姓名、课程名保持一致，按',
    '姓名+学科匹配，不需要再填任教班和周课时（那两项已经在任课表里维护）。',
    '',
    '职务：逗号分隔，比如"班主任"或"备课组长，班主任"。',
    '',
    '固定节次 / 不能排课节次：都用「周X」开头，多个时间段用逗号连接、按"周X"',
    '重新切分（所以数字之间的逗号不会被误当成分段符），比如：',
    '  周一4、5节，周三上午不排课     —— 周一第4、5节 加 周三整个上午',
    '  周五第4，5节不排课             —— 逗号在数字之间，是"第4节、第5节"',
    '  周二上午2、3、4节不排课        —— 上午第2、3、4节（下午同理，第N节=',
    '                                   第5+N节，如"下午2节"是全天第7节）',
    '不写"上午/下午"、只写数字时按全天绝对节次算；不写数字、只写"上午/下午"',
    '时代表一整个半天。固定节次的语义是"窗口"——这门课的周课时会被排在这些',
    '节次里，不是要求全部占满（比如窗口2格、周课时1节，就是2选1）。',
    '',
    '排课要求：目前支持这几种写法（照抄替换数字即可，其余写法解析不了会报错）：',
    '  保证每天有1节                          —— 该学科系每天至少1节',
    '  同一个班当天不能排2节                  —— 该学科系当天最多1节（数字是',
    '                                            "不能排"的那个数，实际上限要',
    '                                            再减1）',
    '  保证周一三四每天1节                    —— 只在写到的这几天每天1节',
    '  与心理课分单双周上，即"心美"周课时1节   —— 美术那一行这么写',
    '  与美术课分单双周上，即"心美"周课时1节   —— 心理那一行这么写（谁单周谁',
    '                                            双周由系统统一分配，不用管）',
    '',
    '备注：目前支持这几种写法，可以写在同一段里：',
    '  其中一天为连堂课                       —— 一周里有一天要连续两节',
    '  两个班之间要隔开1节                    —— 两个班的课之间至少隔1节',
    '',
    '解析结果会在导入预览页面逐条回显，确认没有歧义再提交——如果开启了 AI',
    '复核，两边解析结果不一致的地方也会在预览里标出来，供你核对。',
]


def build_rules_sheet_template() -> 'openpyxl.Workbook':
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '排课说明'
    ws.append(['占位表头行，程序从第3行起读'])
    ws.append(['姓名', '任教年级', '学科', '职务', '固定节次', '不能排课节次', '排课要求', '备注'])
    for row in _RULE_SHEET_EXAMPLES:
        ws.append(list(row))

    notes = wb.create_sheet('填写说明')
    for line in _RULE_SHEET_NOTES:
        notes.append([line])
    notes.column_dimensions['A'].width = 90
    return wb
