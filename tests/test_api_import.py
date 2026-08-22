import io
import shutil
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from scheduler.api.app import app

ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / 'scheduler' / 'config'


def _teaching_table_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['初三'])
    ws.append(['班别', '语文', '数学'])
    ws.append([1, '李琼', '徐仪涵'])
    ws.append([2, '郑艳秀', '徐仪涵'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _rules_sheet_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['占位表头行'])
    ws.append(['姓名', '任教年级', '学科', '任教班', '周课时', '职务',
               '固定节次', '不能排课节次', '排课要求', '备注'])
    ws.append(['李琼', '初三', '语文', '1', 6, None, None, None, '保证每天有1节', None])
    ws.append(['郑艳秀', '初三', '语文', '2', 6, None, None, None, '保证每天有1节', None])
    ws.append(['徐仪涵', '初三', '数学', '1,2', 5, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _rules_sheet_bytes_conflicting():
    """1 班语文在排课说明里写成"王老师"，与任课表的"李琼"对不上——
    唯一制造冲突的地方，其余与 `_rules_sheet_bytes` 保持一致。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['占位表头行'])
    ws.append(['姓名', '任教年级', '学科', '任教班', '周课时', '职务',
               '固定节次', '不能排课节次', '排课要求', '备注'])
    ws.append(['王老师', '初三', '语文', '1', 6, None, None, None, None, None])
    ws.append(['郑艳秀', '初三', '语文', '2', 6, None, None, None, None, None])
    ws.append(['徐仪涵', '初三', '数学', '1,2', 5, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _copy_static_config_into(dest_dir):
    """`load_config` 需要 courses/venues/plans 静态配置；把仓库里真实的这
    三份复制进隔离目录，既让 load_config 能跑通，又不污染真实配置。"""
    for name in ('courses.yaml', 'venues.yaml', 'plans.yaml'):
        shutil.copy(CONFIG_DIR / name, dest_dir / name)


@pytest.fixture()
def client():
    return TestClient(app)


def test_import_returns_preview_with_no_conflicts(client):
    resp = client.post(
        '/api/import',
        params={'grade': '初三'},
        files={
            'teaching_file': ('任课表.xlsx', _teaching_table_bytes(),
                              'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            'rules_file': ('排课说明.xlsx', _rules_sheet_bytes(),
                           'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body['conflicts'] == []
    assert body['classes'] == 2
    assert '保证每天有1节' in [item['raw'] for item in body['rule_echo']['排课要求']]
    assert body['token']


def test_import_confirm_writes_config_on_success(client, tmp_path, monkeypatch):
    import scheduler.api.routes as routes_module
    # DEFAULT_CONFIG_DIR 被换成隔离的 tmp_path，避免测试写脏仓库里的真实
    # scheduler/config；但 load_config 仍要读 courses/venues/plans 静态配置，
    # 所以先把这三份从真实 CONFIG_DIR 复制进 tmp_path。
    _copy_static_config_into(tmp_path)
    monkeypatch.setattr(routes_module, 'DEFAULT_CONFIG_DIR', tmp_path)
    # 先拿一份没有冲突的预览
    resp = client.post(
        '/api/import', params={'grade': '初三'},
        files={
            'teaching_file': ('任课表.xlsx', _teaching_table_bytes(), 'application/octet-stream'),
            'rules_file': ('排课说明.xlsx', _rules_sheet_bytes(), 'application/octet-stream'),
        },
    )
    token = resp.json()['token']
    confirm = client.post('/api/import/confirm', json={'token': token})
    assert confirm.status_code == 200
    assert (tmp_path / 'teaching.yaml').exists()
    assert (tmp_path / 'rules.generated.yaml').exists()


def test_import_confirm_rejects_when_conflicted(client, tmp_path, monkeypatch):
    """铁律：两份文件教师归属对不上必须硬性阻断确认，不允许静默二选一。"""
    import scheduler.api.routes as routes_module
    _copy_static_config_into(tmp_path)
    monkeypatch.setattr(routes_module, 'DEFAULT_CONFIG_DIR', tmp_path)

    resp = client.post(
        '/api/import', params={'grade': '初三'},
        files={
            'teaching_file': ('任课表.xlsx', _teaching_table_bytes(), 'application/octet-stream'),
            'rules_file': ('排课说明.xlsx', _rules_sheet_bytes_conflicting(),
                           'application/octet-stream'),
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body['conflicts']   # 1班语文：任课表说李琼，排课说明说王老师

    confirm = client.post('/api/import/confirm', json={'token': body['token']})
    assert confirm.status_code == 400
    assert not (tmp_path / 'teaching.yaml').exists()
    assert not (tmp_path / 'rules.generated.yaml').exists()


def test_import_confirm_rejects_unknown_token(client):
    resp = client.post('/api/import/confirm', json={'token': '不存在'})
    assert resp.status_code == 404


def test_config_status_reflects_written_config(client, tmp_path, monkeypatch):
    import scheduler.api.routes as routes_module
    monkeypatch.setattr(routes_module, 'DEFAULT_CONFIG_DIR', tmp_path)
    resp = client.get('/api/config/status')
    assert resp.status_code == 200
    assert resp.json()['ready'] is False


def test_import_confines_uploaded_files_despite_path_traversal_filename(client, tmp_path, monkeypatch):
    """`teaching_file.filename`/`rules_file.filename` 是客户端可控字符串。
    把 `tempfile.TemporaryDirectory()` 钉死到一个可检查的目录上，验证恶意
    文件名（`../../../evil.xlsx`）不会让写入逃出这个目录——只应该看到
    固定命名的 teaching.xlsx / rules.xlsx，不该出现任何 evil* 文件。"""
    import scheduler.api.routes as routes_module

    class _FixedTempDir:
        def __enter__(self):
            return str(tmp_path)

        def __exit__(self, *exc_info):
            return False

    monkeypatch.setattr(routes_module.tempfile, 'TemporaryDirectory',
                        lambda: _FixedTempDir())

    resp = client.post(
        '/api/import', params={'grade': '初三'},
        files={
            'teaching_file': ('../../../evil-teaching.xlsx', _teaching_table_bytes(),
                              'application/octet-stream'),
            'rules_file': ('../../../evil-rules.xlsx', _rules_sheet_bytes(),
                           'application/octet-stream'),
        },
    )
    assert resp.status_code == 200
    written = {p.name for p in tmp_path.iterdir()}
    assert written == {'teaching.xlsx', 'rules.xlsx'}
    assert not (tmp_path.parent / 'evil-teaching.xlsx').exists()
    assert not (tmp_path.parent / 'evil-rules.xlsx').exists()


def test_get_plan_returns_400_when_config_missing(client, tmp_path, monkeypatch):
    """`load_config` 缺配置时抛 `ConfigError`——必须被转成格式化的 400，
    不能裸传播成未处理的 500。"""
    import scheduler.api.routes as routes_module
    monkeypatch.setattr(routes_module, 'DEFAULT_CONFIG_DIR', tmp_path)
    resp = client.get('/api/config/plan', params={'grade': '初三'})
    assert resp.status_code == 400
