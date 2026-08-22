from pathlib import Path

import openpyxl
import pytest

from scheduler.core import calendar as cal
from scheduler.core.exporter import export_to_template
from scheduler.core.models import Dataset, Teacher
from scheduler.core.solver import Placement, Solution


def _blank_template(path, sheet_name='下学期'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    # 模拟模板已预填的教务固定占位格——周一第二课堂（第9节）= 班会
    ws.cell(row=12, column=10, value='班')
    ws.cell(row=13, column=10, value='班')
    wb.save(path)


def test_places_course_at_the_templates_row_and_column(tmp_path):
    template = tmp_path / '模板.xlsx'
    _blank_template(template)

    dataset = Dataset(grade='初三', classes=[1, 2],
                      teachers={'张老师': Teacher(name='张老师'), '李老师': Teacher(name='李老师')},
                      tasks=[])
    solution = Solution(status='OPTIMAL', wall_time=0.1, placements=[
        Placement(task_id=0, class_id=1, course='语文', teacher='张老师',
                  slot=cal.slot_index(0, 1)),          # 初三1 周一第1节
        Placement(task_id=1, class_id=2, course='数学', teacher='李老师',
                  slot=cal.slot_index(4, 8)),          # 初三2 周五第8节
    ])

    out = tmp_path / '结果.xlsx'
    export_to_template(solution, dataset, template, out)

    wb = openpyxl.load_workbook(out)
    ws = wb['下学期']
    assert ws.cell(row=12, column=2).value == '语文'
    assert ws.cell(row=13, column=45).value == '数学'


def test_leaves_reserved_prefilled_cells_untouched(tmp_path):
    """模板里教务固定占位的格子已经有内容，导出不该覆盖它。"""
    template = tmp_path / '模板.xlsx'
    _blank_template(template)

    dataset = Dataset(grade='初三', classes=[1],
                      teachers={'张老师': Teacher(name='张老师')}, tasks=[])
    solution = Solution(status='OPTIMAL', wall_time=0.1, placements=[
        Placement(task_id=0, class_id=1, course='语文', teacher='张老师',
                  slot=cal.slot_index(0, 1)),
    ])

    out = tmp_path / '结果.xlsx'
    export_to_template(solution, dataset, template, out)

    wb = openpyxl.load_workbook(out)
    ws = wb['下学期']
    assert ws.cell(row=12, column=10).value == '班'   # 周一第二课堂（班会）原样保留
    assert ws.cell(row=13, column=10).value == '班'


def test_unknown_sheet_name_raises(tmp_path):
    template = tmp_path / '模板.xlsx'
    _blank_template(template)
    dataset = Dataset(grade='初三', classes=[1], teachers={}, tasks=[])
    solution = Solution(status='OPTIMAL', wall_time=0.1, placements=[])
    with pytest.raises(ValueError, match='没有名为'):
        export_to_template(solution, dataset, template, tmp_path / 'x.xlsx',
                           sheet_name='不存在的工作表')
