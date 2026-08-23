from pathlib import Path

import openpyxl
import pytest

from scheduler.core.config import load_config
from scheduler.core.importer import parse_teaching_table

CONFIG_DIR = Path(__file__).resolve().parents[1] / "scheduler" / "config"


@pytest.fixture(scope="module")
def cfg():
    return load_config(CONFIG_DIR)


def _write_teaching_table(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["初三"])
    ws.append(["班别", "语文", "数学"])
    ws.append([1, "李琼", "徐仪涵"])
    ws.append([2, "郑艳秀", "徐仪涵"])
    wb.save(path)


def test_parse_teaching_table_builds_class_course_to_teacher_map(tmp_path, cfg):
    path = tmp_path / "任课表.xlsx"
    _write_teaching_table(path)
    pivot = parse_teaching_table(path, cfg)
    assert pivot[(1, "语文")] == "李琼"
    assert pivot[(1, "数学")] == "徐仪涵"
    assert pivot[(2, "语文")] == "郑艳秀"
    assert pivot[(2, "数学")] == "徐仪涵"
    assert len(pivot) == 4


def test_parse_teaching_table_rejects_unknown_course(tmp_path, cfg):
    path = tmp_path / "任课表.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["初三"])
    ws.append(["班别", "不存在的课"])
    ws.append([1, "某老师"])
    wb.save(path)
    with pytest.raises(ValueError, match="不在课程目录里"):
        parse_teaching_table(path, cfg)


from scheduler.core.importer import ImportResult, merge_teaching_and_rules


def _write_rules_sheet(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["占位表头行，程序从第3行起读"])
    ws.append(["姓名", "任教年级", "学科", "任教班", "周课时", "职务",
               "固定节次", "不能排课节次", "排课要求", "备注"])
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_merge_builds_tasks_when_both_sources_agree(tmp_path, cfg):
    teaching_path = tmp_path / "任课表.xlsx"
    _write_teaching_table(teaching_path)
    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["李琼", "初三", "语文", "1", 6, None, None, None, "保证每天有1节", None],
        ["郑艳秀", "初三", "语文", "2", 6, None, None, None, "保证每天有1节", None],
        ["徐仪涵", "初三", "数学", "1,2", 5, None, None, None, None, None],
    ])
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三")
    assert isinstance(result, ImportResult)
    assert result.conflicts == []
    by_key = {(t.class_id, t.course): t.teacher for t in result.dataset.tasks}
    assert by_key[(1, "语文")] == "李琼"
    assert by_key[(2, "数学")] == "徐仪涵"


def test_merge_flags_teacher_mismatch_as_conflict(tmp_path, cfg):
    teaching_path = tmp_path / "任课表.xlsx"
    _write_teaching_table(teaching_path)   # 1班语文=李琼
    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["王老师", "初三", "语文", "1", 6, None, None, None, None, None],   # 排课说明说是王老师
        ["郑艳秀", "初三", "语文", "2", 6, None, None, None, None, None],   # 2班语文与任课表一致，隔离出唯一的冲突场景
        ["徐仪涵", "初三", "数学", "1,2", 5, None, None, None, None, None],
    ])
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三")
    assert len(result.conflicts) == 1
    conflict = result.conflicts[0]
    assert conflict["class_id"] == 1 and conflict["course"] == "语文"
    assert conflict["from_teaching_table"] == "李琼"
    assert conflict["from_rules_sheet"] == "王老师"
    by_key = {(t.class_id, t.course): t.teacher for t in result.dataset.tasks}
    assert (1, "语文") not in by_key


def test_merge_flags_missing_rules_row_as_conflict(tmp_path, cfg):
    teaching_path = tmp_path / "任课表.xlsx"
    _write_teaching_table(teaching_path)   # 任课表里有 1班数学=徐仪涵
    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["李琼", "初三", "语文", "1", 6, None, None, None, None, None],
        # 数学没有对应行——任课表说 1班数学是徐仪涵，排课说明查不到周课时
    ])
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三")
    missing = [c for c in result.conflicts if c["course"] == "数学"]
    assert len(missing) >= 1
    assert missing[0]["from_teaching_table"] == "徐仪涵"
    assert missing[0]["from_rules_sheet"] is None


def test_merge_builds_rule_echo_for_review(tmp_path, cfg):
    teaching_path = tmp_path / "任课表.xlsx"
    _write_teaching_table(teaching_path)
    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["李琼", "初三", "语文", "1", 6, None, None, "周五上午不排课", "保证每天有1节", None],
        ["徐仪涵", "初三", "数学", "1,2", 5, None, None, None, None, None],
    ])
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三")
    echo = result.rule_echo["不能排课节次"]
    assert any(item["raw"] == "周五上午不排课" for item in echo)
    req_echo = result.rule_echo["排课要求"]
    assert any(item["raw"] == "保证每天有1节" for item in req_echo)


def test_merge_ai_engine_uses_parse_row_ai(tmp_path, cfg):
    import json
    from types import SimpleNamespace

    class FakeMessages:
        def __init__(self):
            self.call_count = 0

        def create(self, **kwargs):
            self.call_count += 1
            payload = {
                "not_available": [], "fixed_slots": [],
                "requirement": [{"type": "daily_min", "params": {"n": 1}}],
                "remark": [],
            }
            return SimpleNamespace(content=[SimpleNamespace(text=json.dumps(payload))])

    class FakeClient:
        def __init__(self):
            self.messages = FakeMessages()

    teaching_path = tmp_path / "任课表.xlsx"
    _write_teaching_table(teaching_path)
    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["李琼", "初三", "语文", "1", 6, None, None, None, "保证每天有1节", None],
        ["徐仪涵", "初三", "数学", "1,2", 5, None, None, None, None, None],
    ])
    client = FakeClient()
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三",
                                      rule_engine="ai", ai_client=client)
    # 关键断言：证明 AI 代码路径真的被调用了（原断言在正则也能推出同一结果的
    # 输入文本下形同虚设——即使 rule_engine="ai" 被静默忽略、实际走了正则，
    # 也会得到同样的 daily_min 片段，测不出问题）。
    assert client.messages.call_count == 2   # 排课说明两行，每行一次 AI 调用
    daily_min = [r for r in result.rules if r["type"] == "daily_min"]
    assert len(daily_min) >= 1


def test_merge_flips_parity_by_class_id_for_half_period_courses(tmp_path, cfg):
    """port 自 import_excel 的坑 3 修复：merge 路径也要按班号奇偶翻转单双周，
    否则同一份 Excel 走两条导入路径会得到不同的负荷分布。"""
    teaching_path = tmp_path / "任课表.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["初三"])
    ws.append(["班别", "美术"])
    for class_id in range(1, 5):
        ws.append([class_id, "赵老师"])
    wb.save(teaching_path)

    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["赵老师", "初三", "美术", "1,2,3,4", 0.5, None, None, None, None, None],
    ])
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三")
    by_class = {t.class_id: t.parity for t in result.dataset.tasks if t.course == "美术"}
    # 课程目录声明美术默认单周；奇数班保持默认，偶数班翻转为双周（见 CLAUDE.md 坑 3）。
    assert by_class[1] == "单周" and by_class[3] == "单周"
    assert by_class[2] == "双周" and by_class[4] == "双周"


def test_merge_ai_engine_rule_echo_keeps_requirement_and_remark_separate(tmp_path, cfg):
    """排课要求与备注在 AI 模式下要各自只回显自己那一列的解析结果，
    不能像修复前那样把 requirement+remark 合并片段同时塞进两个 key。"""
    import json
    from types import SimpleNamespace

    class FakeMessages:
        def create(self, **kwargs):
            payload = {
                "not_available": [], "fixed_slots": [],
                "requirement": [{"type": "daily_min", "params": {"n": 1}}],
                "remark": [{"type": "spacing", "params": {"n": 2}}],
            }
            return SimpleNamespace(content=[SimpleNamespace(text=json.dumps(payload))])

    class FakeClient:
        messages = FakeMessages()

    teaching_path = tmp_path / "任课表.xlsx"
    _write_teaching_table(teaching_path)
    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["李琼", "初三", "语文", "1", 6, None, None, None, "保证每天有1节", "两个班之间要隔开2节"],
        ["徐仪涵", "初三", "数学", "1,2", 5, None, None, None, None, None],
    ])
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三",
                                      rule_engine="ai", ai_client=FakeClient())
    req_echo = result.rule_echo["排课要求"][0]["parsed"]
    remark_echo = result.rule_echo["备注"][0]["parsed"]
    assert "daily_min" in req_echo and "spacing" not in req_echo
    assert "spacing" in remark_echo and "daily_min" not in remark_echo


def test_merge_ai_engine_uses_ai_not_available_for_forbidden_slots(tmp_path, cfg):
    """选 rule_engine='ai' 时，教师禁排应来自 AI 的 not_available，不能悄悄退化成正则。

    "周三上午不排课" 正则会解析成周三上午 5 节的并集；构造一个刻意不同的假 AI
    答案（周五第9节），断言最终 forbidden 是 AI 的答案而不是正则的答案。
    """
    import json
    from types import SimpleNamespace

    class FakeMessages:
        def create(self, **kwargs):
            payload = {
                "not_available": [[4, 9]],
                "fixed_slots": [], "requirement": [], "remark": [],
            }
            return SimpleNamespace(content=[SimpleNamespace(text=json.dumps(payload))])

    class FakeClient:
        messages = FakeMessages()

    teaching_path = tmp_path / "任课表.xlsx"
    _write_teaching_table(teaching_path)
    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["李琼", "初三", "语文", "1", 6, None, None, "周三上午不排课", None, None],
        ["徐仪涵", "初三", "数学", "1,2", 5, None, None, None, None, None],
    ])
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三",
                                      rule_engine="ai", ai_client=FakeClient())
    forbidden = result.dataset.teachers["李琼"].forbidden_slots()
    assert forbidden == {(4, 9)}


def test_merge_ai_engine_rule_echo_shows_ai_not_available_not_regex(tmp_path, cfg):
    """不能排课节次的回显必须和实际写盘的规则一致（AI 模式下来自 AI 的答案），
    否则教务确认的是正则解释，系统实际生效的却是 AI 解释——违反必须回显确认的铁律。

    "周三上午不排课" 正则会解析成周三（day=2）上午 5 节的并集；构造一个刻意不同的
    假 AI 答案（周五第9节，即 day=4, period=9），断言 rule_echo 显示的是 AI 的答案。
    """
    import json
    from types import SimpleNamespace

    class FakeMessages:
        def create(self, **kwargs):
            payload = {
                "not_available": [[4, 9]],
                "fixed_slots": [], "requirement": [], "remark": [],
            }
            return SimpleNamespace(content=[SimpleNamespace(text=json.dumps(payload))])

    class FakeClient:
        messages = FakeMessages()

    teaching_path = tmp_path / "任课表.xlsx"
    _write_teaching_table(teaching_path)
    rules_path = tmp_path / "排课说明.xlsx"
    _write_rules_sheet(rules_path, [
        ["李琼", "初三", "语文", "1", 6, None, None, "周三上午不排课", None, None],
        ["徐仪涵", "初三", "数学", "1,2", 5, None, None, None, None, None],
    ])
    result = merge_teaching_and_rules(teaching_path, rules_path, cfg, grade="初三",
                                      rule_engine="ai", ai_client=FakeClient())
    echo = result.rule_echo["不能排课节次"][0]["parsed"]
    # AI 答案是周五第9节；正则会把"周三上午"解析成周三1-5节，两者互斥，
    # 用来验证回显没有悄悄退化成正则的输出。
    assert "周五" in echo and "9" in echo
    assert "周三" not in echo
