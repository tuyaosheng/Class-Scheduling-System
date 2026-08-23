from collections import Counter, defaultdict
from pathlib import Path

import pytest

from scheduler.core import calendar as cal
from scheduler.core.config import load_config
from scheduler.core.importer import import_excel
from scheduler.core.rules import load_rules
from scheduler.core.solver import solve, solve_many

ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / 'scheduler' / 'config'
EXCEL = ROOT / '任课与排课说明.xlsx'


@pytest.fixture(scope='module')
def solved():
    cfg = load_config(CONFIG_DIR)
    result = import_excel(EXCEL, cfg, grade='初三')
    rules = load_rules(CONFIG_DIR / 'rules.yaml', CONFIG_DIR / 'rules.generated.yaml')
    solution = solve(result.dataset, cfg, rules, max_seconds=30)
    return cfg, result.dataset, solution


def test_grade3_is_feasible(solved):
    _, _, solution = solved
    assert solution.feasible, '初三 32 班应可排，状态=%s' % solution.status


def test_solve_time_is_interactive(solved):
    """设计文档实测 0.6 秒。明显慢于此说明建模有问题，不要靠调大超时掩盖。"""
    _, _, solution = solved
    assert solution.wall_time < 5.0, '耗时 %.1fs，远超基线 0.6s' % solution.wall_time


def test_placement_count_matches_total_periods(solved):
    _, dataset, solution = solved
    assert len(solution.placements) == sum(t.periods for t in dataset.tasks)


def test_every_class_uses_37_distinct_slots(solved):
    _, dataset, solution = solved
    by_class = defaultdict(set)
    for p in solution.placements:
        if p.parity != '双周':
            by_class[p.class_id].add(p.slot)
    assert set(Counter({c: len(s) for c, s in by_class.items()}).values()) == {37}


def test_no_class_double_booked(solved):
    _, _, solution = solved
    for parity in ('单周', '双周'):
        seen = Counter((p.class_id, p.slot) for p in solution.placements
                       if p.parity in (None, parity))
        assert max(seen.values()) == 1


def test_no_teacher_double_booked(solved):
    _, _, solution = solved
    for parity in ('单周', '双周'):
        seen = Counter((p.teacher, p.slot) for p in solution.placements
                       if p.parity in (None, parity))
        assert max(seen.values()) == 1


def test_art_and_psych_share_slots(solved):
    _, _, solution = solved
    art = {(p.class_id, p.slot) for p in solution.placements if p.course == '美术'}
    psy = {(p.class_id, p.slot) for p in solution.placements if p.course == '心理'}
    assert art == psy and len(art) == 32


def test_reserved_slots_stay_empty_in_solver_output(solved):
    """班会/体比/体选/校本1/综实2 是教务固定安排，求解器不该在这 8 格放任何课。"""
    cfg, _, solution = solved
    reserved = cfg.reserved_slot_indices('初三')
    assert not any(p.slot in reserved for p in solution.placements)


def test_physics_family_every_day(solved):
    cfg, _, solution = solved
    by_class_day = defaultdict(int)
    for p in solution.placements:
        if cfg.family_of(p.course) == '物理':
            by_class_day[(p.class_id, cal.slot_of(p.slot)[0])] += 1
    for class_id in range(1, 33):
        for day in range(5):
            assert by_class_day[(class_id, day)] >= 1, '%d班周%d没有物理系课程' % (class_id, day + 1)


def test_export_excel(tmp_path, solved):
    import openpyxl
    from scheduler.core.exporter import export_excel
    cfg, dataset, solution = solved
    path = tmp_path / '课表.xlsx'
    export_excel(solution, dataset, path, cfg=cfg)
    wb = openpyxl.load_workbook(path)
    assert '班级课表' in wb.sheetnames and '教师课表' in wb.sheetnames
    ws = wb['班级课表']
    assert ws.max_row == cal.N_SLOTS + 1        # 表头 + 45 格
    assert ws.max_column == 32 + 2              # 星期、节次 + 32 个班
    # 周一第9节（班会，教务固定安排）对每个班都应显示占位标签，而非空白
    row = cal.slot_index(0, 9) + 2
    assert all(ws.cell(row=row, column=col).value == '（教务固定安排）'
               for col in range(3, 3 + 32))


def test_cli_solve(tmp_path, monkeypatch, capsys):
    from scheduler.cli import main
    monkeypatch.chdir(ROOT)
    out = tmp_path / '课表.xlsx'
    assert main(['solve', '--out', str(out), '--max-seconds', '30']) == 0
    assert out.exists()
    out_text = capsys.readouterr().out
    assert 'OPTIMAL' in out_text or 'FEASIBLE' in out_text


# ---------------------------------------------------------------- 多解生成

def test_solve_many_returns_distinct_feasible_solutions():
    cfg = load_config(CONFIG_DIR)
    result = import_excel(EXCEL, cfg, grade='初三')
    rules = load_rules(CONFIG_DIR / 'rules.yaml', CONFIG_DIR / 'rules.generated.yaml')
    solutions = solve_many(result.dataset, cfg, rules, count=3, min_diff=8, max_seconds=30)
    assert len(solutions) == 3
    for sol in solutions:
        assert sol.feasible

    def placed(sol):
        return {(p.task_id, p.slot) for p in sol.placements}

    for i in range(len(solutions)):
        for j in range(i + 1, len(solutions)):
            diff = placed(solutions[i]) ^ placed(solutions[j])
            assert len(diff) >= 8, '第 %d 与第 %d 个解差异只有 %d 处' % (i, j, len(diff))


def test_solve_many_each_solution_passes_verification():
    from scheduler.core.verifier import verify
    cfg = load_config(CONFIG_DIR)
    result = import_excel(EXCEL, cfg, grade='初三')
    rules = load_rules(CONFIG_DIR / 'rules.yaml', CONFIG_DIR / 'rules.generated.yaml')
    solutions = solve_many(result.dataset, cfg, rules, count=2, min_diff=8, max_seconds=30)
    for sol in solutions:
        violations = verify(sol, result.dataset, cfg, rules)
        assert violations == [], '\n'.join(v.detail for v in violations)


def test_solve_many_stops_early_when_diversity_exhausted():
    """可行解总数比要求的 count 少时，应少给几个而不是报错或死等。

    用一个把域收窄到 3 格、放 2 节课的最小场景：C(3,2)=3 种放法是
    这个问题的解空间上限，要 5 个必然只能拿到 3 个。
    """
    from scheduler.core.models import Dataset, Teacher, TeachingTask
    from scheduler.core.rules import Rule

    cfg = load_config(CONFIG_DIR)
    task = TeachingTask(id=0, grade='初三', class_id=1, course='语文',
                        teacher='张老师', periods=2)
    dataset = Dataset(grade='初三', classes=[1],
                      teachers={'张老师': Teacher(name='张老师')}, tasks=[task])
    allowed = {cal.slot_index(0, 1), cal.slot_index(0, 2), cal.slot_index(0, 3)}
    forbidden = [list(cal.slot_of(s)) for s in range(cal.N_SLOTS) if s not in allowed]
    rules = [Rule(type='forbid_slots', scope={'course': '语文'}, params={'slots': forbidden})]

    solutions = solve_many(dataset, cfg, rules, count=5, min_diff=1, max_seconds=5)
    assert len(solutions) == 3
